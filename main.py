import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict
import csv
import os

# Constants
DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
START_TIME = time(9, 0)
END_TIME = time(18, 30)
LECTURE_DURATION = 3  # 1.5 hours
LAB_DURATION = 4      # 2 hours
TUTORIAL_DURATION = 2 # 1 hour
SELF_STUDY_DURATION = 2 # 1 hour
BREAK_DURATION = 1    # 0.5 hour

LUNCH_WINDOW_START = time(12, 30)
LUNCH_WINDOW_END = time(14, 0)
LUNCH_DURATION = 60

# Visually appealing color palette
COLOR_PALETTE = [
    "A8D5BA",  # Soft Green
    "FFDAB9",  # Peach
    "B0C4DE",  # Light Steel Blue
    "FFFACD",  # Lemon Chiffon
    "D8BFD8",  # Thistle
    "F0E68C",  # Khaki
    "E6E6FA",  # Lavender
    "98FB98",  # Pale Green
    "FFDEAD",  # Navajo White
    "ADD8E6"   # Light Blue
]

basket_group_colors = {
    'B1': "FF9999",  # Soft Coral
    'B2': "99FFCC",  # Light Mint
    'B3': "99CCFF",  # Pale Blue
    'B4': "FFCC99"   # Warm Yellow
}

# Fixed slots for basket electives
BASKET_SLOTS = {
    'B1': 0,   # 9:00-10:30
    'B2': 3,   # 10:30-12:00
    'B3': 7,   # 14:00-15:30 (post-lunch)
    'B4': 10   # 15:30-17:00
}

TIME_SLOTS = []
lunch_breaks = {}

def generate_time_slots():
    slots = []
    current_time = datetime.combine(datetime.today(), START_TIME)
    end_time = datetime.combine(datetime.today(), END_TIME)
    while current_time < end_time:
        current = current_time.time()
        next_time = (current_time + timedelta(minutes=30)).time()
        slots.append((current, next_time))
        current_time += timedelta(minutes=30)
    return slots

def calculate_lunch_breaks(semesters):
    global lunch_breaks
    lunch_breaks = {}
    total_semesters = len(semesters)
    if total_semesters == 0:
        return
    total_window_minutes = (LUNCH_WINDOW_END.hour * 60 + LUNCH_WINDOW_END.minute -
                           LUNCH_WINDOW_START.hour * 60 - LUNCH_WINDOW_START.minute)
    stagger_interval = (total_window_minutes - LUNCH_DURATION) / (total_semesters - 1) if total_semesters > 1 else 0
    sorted_semesters = sorted(semesters)
    for i, semester in enumerate(sorted_semesters):
        start_minutes = LUNCH_WINDOW_START.hour * 60 + LUNCH_WINDOW_START.minute + int(i * stagger_interval)
        start_hour, start_min = divmod(start_minutes, 60)
        end_minutes = start_minutes + LUNCH_DURATION
        end_hour, end_min = divmod(end_minutes, 60)
        lunch_breaks[semester] = (time(start_hour, start_min), time(end_hour, end_min))

def load_rooms():
    rooms = {}
    default_rooms = {
        'R1': {'capacity': 70, 'type': 'LECTURE_ROOM', 'roomNumber': 'R101', 'schedule': {day: set() for day in range(len(DAYS))}},
        'R2': {'capacity': 70, 'type': 'LECTURE_ROOM', 'roomNumber': 'R102', 'schedule': {day: set() for day in range(len(DAYS))}},
        'L1': {'capacity': 35, 'type': 'COMPUTER_LAB', 'roomNumber': 'L101', 'schedule': {day: set() for day in range(len(DAYS))}},
        'S1': {'capacity': 120, 'type': 'SEATER_120', 'roomNumber': 'S101', 'schedule': {day: set() for day in range(len(DAYS))}}
    }
    try:
        with open('classrooms.csv', 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rooms[row['id']] = {
                    'capacity': int(row['capacity']),
                    'type': row['type'],
                    'roomNumber': row['roomNumber'],
                    'schedule': {day: set() for day in range(len(DAYS))}
                }
    except FileNotFoundError:
        print("Warning: classrooms.csv not found, using default rooms")
        return default_rooms
    except Exception as e:
        print(f"Warning: Error reading classrooms.csv: {e}, using default rooms")
        return default_rooms
    return rooms

def load_batch_data(df):
    batch_info = {}
    default_batch = {
        ('CSE', 3): {'total': 140, 'num_sections': 2, 'section_size': 70},
        ('ECE', 3): {'total': 70, 'num_sections': 1, 'section_size': 70},
        ('ELECTIVE', 'B1-001'): {'total': 35, 'num_sections': 1, 'section_size': 35},
        ('ELECTIVE', 'B1-002'): {'total': 35, 'num_sections': 1, 'section_size': 35},
        ('ELECTIVE', 'B2-001'): {'total': 35, 'num_sections': 1, 'section_size': 35},
        ('ELECTIVE', 'B3-001'): {'total': 35, 'num_sections': 1, 'section_size': 35},
        ('ELECTIVE', 'B4-001'): {'total': 35, 'num_sections': 1, 'section_size': 35}
    }
    if df is None or df.empty:
        print("Warning: No course data provided, using default batch info")
        return default_batch
    try:
        grouped = df.groupby(['Department', 'Semester'])
        for (dept, sem), group in grouped:
            if 'total_students' in group.columns and not group['total_students'].isna().all():
                total_students = int(group['total_students'].max())
                max_batch_size = 70
                num_sections = (total_students + max_batch_size - 1) // max_batch_size
                section_size = (total_students + num_sections - 1) // num_sections
                batch_info[(dept, sem)] = {'total': total_students, 'num_sections': num_sections, 'section_size': section_size}
        basket_courses = df[df['Course Code'].astype(str).str.contains('^B[0-9]')]
        for basket_group in ['B1', 'B2', 'B3', 'B4']:
            basket_courses_group = basket_courses[basket_courses['Course Code'].astype(str).str.startswith(basket_group)]
            total_students = sum(int(course['total_students']) for _, course in basket_courses_group.iterrows() if 'total_students' in df.columns and pd.notna(course['total_students']))
            if total_students > 0:
                batch_info[('ELECTIVE', basket_group)] = {'total': total_students, 'num_sections': 1, 'section_size': total_students}
            for _, course in basket_courses_group.iterrows():
                code = str(course['Course Code'])
                total_students = int(course['total_students']) if 'total_students' in df.columns and pd.notna(course['total_students']) else 35
                batch_info[('ELECTIVE', code)] = {'total': total_students, 'num_sections': 1, 'section_size': total_students}
    except Exception as e:
        print(f"Warning: Error processing batch data: {e}, using default batch info")
        return default_batch
    return batch_info

def is_basket_course(code):
    return str(code).startswith('B') and '-' in str(code)

def get_basket_group(code):
    if is_basket_course(code):
        return str(code).split('-')[0]
    return None

def find_suitable_room(course_type, department, semester, day, start_slot, duration, rooms, batch_info, timetable, course_code="", used_rooms=None):
    if not rooms:
        return "ROOM"
    required_capacity = 60
    is_basket = is_basket_course(course_code)
    total_students = None
    try:
        if course_code and 'df' in globals():
            course_row = df[df['Course Code'] == course_code]
            if not course_row.empty and 'total_students' in course_row.columns and pd.notna(course_row['total_students'].iloc[0]):
                total_students = int(course_row['total_students'].iloc[0])
            elif is_basket:
                basket_group = get_basket_group(course_code)
                elective_info = batch_info.get(('ELECTIVE', basket_group))
                if elective_info:
                    total_students = elective_info['section_size']
    except Exception as e:
        print(f"Warning: Error retrieving total_students: {e}")
    if total_students:
        required_capacity = total_students
    elif batch_info:
        if is_basket:
            elective_info = batch_info.get(('ELECTIVE', course_code))
            if elective_info:
                required_capacity = elective_info['section_size']
        else:
            dept_info = batch_info.get((department, semester))
            if dept_info:
                required_capacity = dept_info['section_size']
    used_room_ids = set(used_rooms or [])
    if course_type in ['LEC', 'TUT', 'SS'] and required_capacity > 70:
        seater_rooms = {rid: r for rid, r in rooms.items() if 'SEATER' in r['type'].upper()}
        room_id = try_room_allocation(seater_rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids)
        if room_id:
            return room_id
    lecture_rooms = {rid: r for rid, r in rooms.items() if 'LECTURE_ROOM' in r['type'].upper()}
    if is_basket:
        basket_group = get_basket_group(course_code)
        basket_used_rooms = set()
        for slot in range(start_slot, start_slot + duration):
            if slot in timetable[day]:
                slot_data = timetable[day][slot]
                if slot_data['classroom'] and slot_data['type'] is not None:
                    slot_code = slot_data.get('code', '')
                    if get_basket_group(slot_code) != basket_group:
                        basket_used_rooms.add(slot_data['classroom'])
        room_id = try_room_allocation(lecture_rooms, course_type, required_capacity, day, start_slot, duration, basket_used_rooms)
        return room_id or "ROOM"
    return try_room_allocation(lecture_rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids) or "ROOM"

def try_room_allocation(rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids):
    for room_id, room in rooms.items():
        if room_id in used_room_ids or room['capacity'] < required_capacity:
            continue
        slots_free = True
        for i in range(duration):
            if start_slot + i in room['schedule'][day]:
                slots_free = False
                break
        if slots_free:
            for i in range(duration):
                room['schedule'][day].add(start_slot + i)
            return room_id
    return None

def get_required_room_type(course):
    code = str(course.get('Course Code', '')).upper()
    if course.get('P', 0) > 0:
        if 'CS' in code or 'DS' in code:
            return 'COMPUTER_LAB'
        elif 'EC' in code:
            return 'HARDWARE_LAB'
        return 'COMPUTER_LAB'
    return 'LECTURE_ROOM'

def calculate_required_slots(course):
    l = float(course.get('L', 0)) if pd.notna(course.get('L')) else 0
    t = int(course.get('T', 0)) if pd.notna(course.get('T')) else 0
    p = int(course.get('P', 0)) if pd.notna(course.get('P')) else 0
    s = int(course.get('S', 0)) if pd.notna(course.get('S')) else 0
    if s > 0 and l == 0 and t == 0 and p == 0:
        return 0, 0, 0, 0
    lecture_sessions = max(1, round(l * 2 / 3)) if l > 0 else 0
    tutorial_sessions = t
    lab_sessions = p // 2
    self_study_sessions = s // 4 if (l > 0 or t > 0 or p > 0) else 0
    return lecture_sessions, tutorial_sessions, lab_sessions, self_study_sessions

class UnscheduledComponent:
    def __init__(self, department, semester, code, name, faculty, component_type, sessions, section='', reason=''):
        self.department = department
        self.semester = semester
        self.code = code
        self.name = name
        self.faculty = faculty
        self.component_type = component_type
        self.sessions = sessions
        self.section = section
        self.reason = reason

    def __eq__(self, other):
        return isinstance(other, UnscheduledComponent) and (self.department, self.semester, self.code, self.component_type, self.section) == (other.department, other.semester, other.code, other.component_type, other.section)

    def __hash__(self):
        return hash((self.department, self.semester, self.code, self.component_type, self.section))

def is_break_time(slot, semester):
    if semester not in lunch_breaks:
        return False
    start, end = slot
    lunch_start, lunch_end = lunch_breaks[semester]
    return lunch_start <= start < lunch_end

def generate_timetable():
    global TIME_SLOTS, df
    TIME_SLOTS = generate_time_slots()
    rooms = load_rooms()
    try:
        df = pd.read_csv('courses.csv')
        df = df.fillna({'L': 0, 'T': 0, 'P': 0, 'S': 0, 'total_students': 60, 'Course Code': 'UNKNOWN', 'Course Name': 'Unknown Course', 'Faculty': 'Unknown Faculty'})
    except FileNotFoundError:
        print("Warning: courses.csv not found, using default course data")
        df = pd.DataFrame([            {'Department': 'CSE', 'Semester': 3, 'Course Code': 'CS101', 'Course Name': 'Intro to CS', 'Faculty': 'Prof A', 'L': 3, 'T': 1, 'P': 2, 'S': 0, 'total_students': 140},
            {'Department': 'ECE', 'Semester': 3, 'Course Code': 'EC101', 'Course Name': 'Electronics', 'Faculty': 'Prof B', 'L': 3, 'T': 0, 'P': 2, 'S': 0, 'total_students': 70},
            {'Department': 'CSE', 'Semester': 3, 'Course Code': 'B1-001', 'Course Name': 'Elective 1A', 'Faculty': 'Prof C', 'L': 2, 'T': 0, 'P': 0, 'S': 0, 'total_students': 35},
            {'Department': 'ECE', 'Semester': 3, 'Course Code': 'B1-002', 'Course Name': 'Elective 1B', 'Faculty': 'Prof D', 'L': 2, 'T': 0, 'P': 0, 'S': 0, 'total_students': 35},
            {'Department': 'CSE', 'Semester': 3, 'Course Code': 'B2-001', 'Course Name': 'Elective 2A', 'Faculty': 'Prof E', 'L': 2, 'T': 0, 'P': 0, 'S': 0, 'total_students': 35},
            {'Department': 'CSE', 'Semester': 3, 'Course Code': 'B3-001', 'Course Name': 'Elective 3A', 'Faculty': 'Prof F', 'L': 2, 'T': 0, 'P': 0, 'S': 0, 'total_students': 35},
            {'Department': 'ECE', 'Semester': 3, 'Course Code': 'B4-001', 'Course Name': 'Elective 4A', 'Faculty': 'Prof G', 'L': 2, 'T': 0, 'P': 0, 'S': 0, 'total_students': 35}
        ])
    batch_info = load_batch_data(df)
    wb = Workbook()
    wb.remove(wb.active)
    professor_schedule = defaultdict(lambda: {day: set() for day in range(len(DAYS))})
    unscheduled_components = set()
    self_study_courses = []
    departments = df['Department'].unique()
    for department in departments:
        semesters = df[df['Department'] == department]['Semester'].unique()
        calculate_lunch_breaks(semesters)
        for semester in semesters:
            courses = df[(df['Department'] == department) & (df['Semester'] == semester)]
            dept_info = batch_info.get((department, semester), {'num_sections': 1})
            num_sections = dept_info['num_sections']
            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65+section)}"
                ws = wb.create_sheet(title=section_title)
                timetable = {day: {slot: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} for slot in range(len(TIME_SLOTS))} for day in range(len(DAYS))}
                subject_color_map = {}
                course_faculty_map = {}
                color_idx = 0
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    if code not in subject_color_map and code != 'nan':
                        if is_basket_course(code):
                            group = get_basket_group(code)
                            subject_color_map[code] = basket_group_colors.get(group, COLOR_PALETTE[color_idx % len(COLOR_PALETTE)])
                        else:
                            subject_color_map[code] = COLOR_PALETTE[color_idx % len(COLOR_PALETTE)]
                        course_faculty_map[code] = {'name': str(course['Course Name']), 'faculty': str(course['Faculty'])}
                        color_idx += 1
                # Schedule basket electives first
                basket_courses = courses[courses['Course Code'].astype(str).str.contains('^B[0-9]')]
                for basket_group in ['B1', 'B2', 'B3', 'B4']:
                    basket_group_courses = basket_courses[basket_courses['Course Code'].astype(str).str.startswith(basket_group)]
                    if basket_group_courses.empty:
                        continue
                    start_slot = BASKET_SLOTS[basket_group]
                    duration = LECTURE_DURATION
                    for day in range(len(DAYS)):
                        if any(is_break_time(TIME_SLOTS[start_slot + i], semester) for i in range(duration)):
                            continue
                        if any(timetable[day][start_slot + i]['type'] for i in range(duration)):
                            continue
                        used_rooms = set()
                        for _, course in basket_group_courses.iterrows():
                            code = str(course['Course Code'])
                            name = str(course['Course Name'])
                            faculty = str(course['Faculty'])
                            l, t, p, s = calculate_required_slots(course)
                            if l == 0:
                                continue
                            if any(start_slot + i in professor_schedule[faculty][day] for i in range(duration)):
                                unscheduled_components.add(UnscheduledComponent(department, semester, code, name, faculty, 'LEC', 1, section, "Faculty conflict in basket slot"))
                                continue
                            room_id = find_suitable_room('LECTURE_ROOM', department, semester, day, start_slot, duration, rooms, batch_info, timetable, code, used_rooms)
                            if not room_id:
                                unscheduled_components.add(UnscheduledComponent(department, semester, code, name, faculty, 'LEC', 1, section, "No suitable room for basket elective"))
                                continue
                            used_rooms.add(room_id)
                            for i in range(duration):
                                professor_schedule[faculty][day].add(start_slot + i)
                                timetable[day][start_slot + i]['type'] = 'LEC'
                                timetable[day][start_slot + i]['code'] = code if i == 0 else ''
                                timetable[day][start_slot + i]['name'] = name if i == 0 else ''
                                timetable[day][start_slot + i]['faculty'] = faculty if i == 0 else ''
                                timetable[day][start_slot + i]['classroom'] = room_id if i == 0 else ''
                # Schedule non-basket courses
                non_basket_courses = courses[~courses['Course Code'].astype(str).str.contains('^B[0-9]')]
                for _, course in non_basket_courses.iterrows():
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    faculty = str(course['Faculty'])
                    l, t, p, s = calculate_required_slots(course)
                    if s > 0 and l == 0 and t == 0 and p == 0:
                        self_study_courses.append({'code': code, 'name': name, 'faculty': faculty, 'department': department, 'semester': semester})
                        continue
                    for session_type, count, duration in [('LEC', l, LECTURE_DURATION), ('TUT', t, TUTORIAL_DURATION), ('LAB', p, LAB_DURATION), ('SS', s, SELF_STUDY_DURATION)]:
                        for _ in range(count):
                            scheduled = False
                            for day in random.sample(range(len(DAYS)), len(DAYS)):
                                for start_slot in range(len(TIME_SLOTS) - duration + 1):
                                    if any(is_break_time(TIME_SLOTS[start_slot + i], semester) for i in range(duration)):
                                        continue
                                    if any(start_slot + i in professor_schedule[faculty][day] or timetable[day][start_slot + i]['type'] for i in range(duration)):
                                        continue
                                    room_type = get_required_room_type(course) if session_type == 'LAB' else 'LECTURE_ROOM'
                                    room_id = find_suitable_room(room_type, department, semester, day, start_slot, duration, rooms, batch_info, timetable, code)
                                    if room_id:
                                        for i in range(duration):
                                            professor_schedule[faculty][day].add(start_slot + i)
                                            timetable[day][start_slot + i]['type'] = session_type
                                            timetable[day][start_slot + i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot + i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot + i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot + i]['classroom'] = room_id if i == 0 else ''
                                        scheduled = True
                                        break
                                if scheduled:
                                    break
                            if not scheduled:
                                unscheduled_components.add(UnscheduledComponent(department, semester, code, name, faculty, session_type, 1, section, "No suitable slot or room found"))
                # Write timetable to worksheet
                header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
                ws.append(header)
                header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                lec_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                lab_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
                tut_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                ss_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                break_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
                for day_idx, day in enumerate(DAYS):
                    row_num = day_idx + 2
                    ws.append([day])
                    merge_ranges = []
                    for slot_idx in range(len(TIME_SLOTS)):
                        cell_value = ''
                        cell_fill = None
                        if is_break_time(TIME_SLOTS[slot_idx], semester):
                            cell_value = "BREAK"
                            cell_fill = break_fill
                        elif timetable[day_idx][slot_idx]['type']:
                            activity_type = timetable[day_idx][slot_idx]['type']
                            code = timetable[day_idx][slot_idx]['code']
                            classroom = timetable[day_idx][slot_idx]['classroom']
                            faculty = timetable[day_idx][slot_idx]['faculty']
                            if code:
                                duration = {'LEC': LECTURE_DURATION, 'LAB': LAB_DURATION, 'TUT': TUTORIAL_DURATION, 'SS': SELF_STUDY_DURATION}.get(activity_type, 1)
                                cell_fill = PatternFill(start_color=subject_color_map.get(code, "E6E6FA"), end_color=subject_color_map.get(code, "E6E6FA"), fill_type="solid")
                                cell_value = f"{code} {activity_type}\n{classroom}\n{faculty}"
                                if duration > 1:
                                    start_col = get_column_letter(slot_idx + 2)
                                    end_col = get_column_letter(slot_idx + duration + 1)
                                    merge_range = f"{start_col}{row_num}:{end_col}{row_num}"
                                    merge_ranges.append((merge_range, cell_fill))
                        cell = ws.cell(row=row_num, column=slot_idx+2, value=cell_value)
                        if cell_fill:
                            cell.fill = cell_fill
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    for merge_range, fill in merge_ranges:
                        ws.merge_cells(merge_range)
                        merged_cell = ws[merge_range.split(':')[0]]
                        merged_cell.fill = fill
                        merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    ws.cell(row=row_num, column=1).border = border
                    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')
                for col_idx in range(1, len(TIME_SLOTS)+2):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 15
                for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
                    ws.row_dimensions[row[0].row].height = 40
                current_row = len(DAYS) + 4
                if self_study_courses:
                    ws.cell(row=current_row, column=1, value="Self-Study Only Courses").font = Font(bold=True)
                    current_row += 1
                    headers = ['Course Code', 'Course Name', 'Faculty']
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=current_row, column=col, value=header).font = Font(bold=True)
                    current_row += 1
                    for course in self_study_courses:
                        if course['department'] == department and course['semester'] == semester:
                            ws.cell(row=current_row, column=1, value=course['code'])
                            ws.cell(row=current_row, column=2, value=course['name'])
                            ws.cell(row=current_row, column=3, value=course['faculty'])
                            current_row += 1
                    current_row += 2
                dept_unscheduled = [c for c in unscheduled_components if c.department == department and c.semester == semester and (c.section == section if num_sections > 1 else True)]
                if dept_unscheduled:
                    ws.cell(row=current_row, column=1, value="Unscheduled Components").font = Font(bold=True, size=12, color="FF0000")
                    current_row += 2
                    headers = ['Course Code', 'Course Name', 'Faculty', 'Component', 'Sessions', 'Reason']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.border = border
                        cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        ws.column_dimensions[get_column_letter(col)].width = 20
                    current_row += 1
                    for comp in dept_unscheduled:
                        cells = [
                            (comp.code, None),
                            (comp.name, None),
                            (comp.faculty, None),
                            (comp.component_type, None),
                            (comp.sessions, None),
                            (comp.reason or "Could not find suitable slot", None)
                        ]
                        for col, (value, fill) in enumerate(cells, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        current_row += 1
                    current_row += 2
                ws.cell(row=current_row, column=1, value="Legend").font = Font(bold=True, size=12)
                current_row += 2
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 40
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 15
                legend_headers = ['Subject Code', 'Subject Name', 'Faculty', 'Color']
                for col, header in enumerate(legend_headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
                for code, color in subject_color_map.items():
                    if code in course_faculty_map:
                        ws.row_dimensions[current_row].height = 25
                        cells = [
                            (code, None),
                            (course_faculty_map[code]['name'], None),
                            (course_faculty_map[code]['faculty'], None),
                            ('', PatternFill(start_color=color, end_color=color, fill_type="solid"))
                        ]
                        for col, (value, fill) in enumerate(cells, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            if fill:
                                cell.fill = fill
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        current_row += 1
    wb.save('timetable.xlsx')
    print("Generated timetable.xlsx")

if __name__ == "__main__":
    generate_timetable()
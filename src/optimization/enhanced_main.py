import pandas as pd
import random
import json
import logging
import time
from datetime import datetime, time as dt_time, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict
import csv
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Tuple, Optional, Set
import multiprocessing as mp

class TimetableConfig:
    """Configuration manager for timetable settings"""
    
    def __init__(self, config_file: str = 'config.json'):
        self.config = self.load_config(config_file)
        self.setup_logging()
        
    def load_config(self, config_file: str) -> dict:
        """Load configuration from JSON file"""
        try:
            with open(config_file, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"Warning: {config_file} not found, using default configuration")
            return self.get_default_config()
        except json.JSONDecodeError as e:
            print(f"Error parsing {config_file}: {e}, using default configuration")
            return self.get_default_config()
    
    def get_default_config(self) -> dict:
        """Return default configuration"""
        return {
            "timetable_settings": {
                "days": ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"],
                "start_time": "09:00",
                "end_time": "18:30",
                "slot_duration_minutes": 30
            },
            "course_durations": {
                "lecture_duration_slots": 3,
                "lab_duration_slots": 4,
                "tutorial_duration_slots": 2,
                "self_study_duration_slots": 2
            },
            "scheduling": {
                "max_retry_attempts": 10,
                "retry_with_different_seeds": True
            },
            "optimization": {
                "enable_parallel_processing": True,
                "max_workers": 4
            }
        }
    
    def setup_logging(self):
        """Setup logging configuration"""
        if self.config.get('logging', {}).get('enable_logging', False):
            log_level = getattr(logging, self.config['logging'].get('log_level', 'INFO'))
            logging.basicConfig(
                level=log_level,
                format='%(asctime)s - %(levelname)s - %(message)s',
                handlers=[
                    logging.FileHandler(self.config['logging'].get('log_file', 'timetable.log')),
                    logging.StreamHandler()
                ]
            )
        else:
            logging.disable(logging.CRITICAL)

class ConflictResolver:
    """Handles conflict resolution and alternative slot suggestions"""
    
    def __init__(self, config: TimetableConfig):
        self.config = config
        self.alternative_slots_cache = {}
    
    def find_alternative_slots(self, course_type: str, duration: int, 
                             available_slots: List[Tuple[int, int]], 
                             max_suggestions: int = 5) -> List[Tuple[int, int]]:
        """Find alternative time slots for a course"""
        cache_key = (course_type, duration, tuple(available_slots))
        if cache_key in self.alternative_slots_cache:
            return self.alternative_slots_cache[cache_key]
        
        suggestions = []
        for day, start_slot in available_slots:
            if len(suggestions) >= max_suggestions:
                break
            suggestions.append((day, start_slot))
        
        self.alternative_slots_cache[cache_key] = suggestions
        return suggestions
    
    def suggest_conflict_resolution(self, conflicts: List[dict]) -> dict:
        """Suggest resolutions for scheduling conflicts"""
        suggestions = {
            'faculty_conflicts': [],
            'room_conflicts': [],
            'time_conflicts': [],
            'alternative_slots': []
        }
        
        for conflict in conflicts:
            if conflict['type'] == 'faculty_conflict':
                suggestions['faculty_conflicts'].append({
                    'course': conflict['course'],
                    'faculty': conflict['faculty'],
                    'suggestion': f"Consider rescheduling {conflict['course']} to avoid faculty {conflict['faculty']} conflict"
                })
            elif conflict['type'] == 'room_conflict':
                suggestions['room_conflicts'].append({
                    'course': conflict['course'],
                    'room': conflict['room'],
                    'suggestion': f"Try alternative room for {conflict['course']} or reschedule time slot"
                })
        
        return suggestions

class OptimizedScheduler:
    """Enhanced scheduler with optimization and conflict resolution"""
    
    def __init__(self, config: TimetableConfig):
        self.config = config
        self.conflict_resolver = ConflictResolver(config)
        self.performance_stats = {
            'generation_time': 0,
            'conflicts_resolved': 0,
            'retry_attempts': 0,
            'courses_scheduled': 0,
            'courses_unscheduled': 0
        }
    
    def generate_timetable_optimized(self, df: pd.DataFrame, rooms: dict) -> dict:
        """Generate timetable with optimization and conflict resolution"""
        start_time = time.time()
        
        # Extract settings from config
        settings = self.config.config
        days = settings['timetable_settings']['days']
        max_retries = settings['scheduling']['max_retry_attempts']
        enable_parallel = settings['optimization']['enable_parallel_processing']
        
        best_result = None
        best_score = float('inf')
        
        for attempt in range(max_retries):
            logging.info(f"Timetable generation attempt {attempt + 1}/{max_retries}")
            
            # Use different random seed for each attempt
            random.seed(42 + attempt)
            
            try:
                result = self._generate_single_attempt(df, rooms, days, enable_parallel)
                score = self._calculate_schedule_score(result)
                
                if score < best_score:
                    best_result = result
                    best_score = score
                
                self.performance_stats['retry_attempts'] = attempt + 1
                
                # If we have a good enough schedule, break early
                if score == 0:  # No conflicts
                    break
                    
            except Exception as e:
                logging.error(f"Attempt {attempt + 1} failed: {e}")
                continue
        
        self.performance_stats['generation_time'] = time.time() - start_time
        logging.info(f"Generation completed in {self.performance_stats['generation_time']:.2f} seconds")
        
        return best_result or {}
    
    def _generate_single_attempt(self, df: pd.DataFrame, rooms: dict, 
                               days: List[str], enable_parallel: bool) -> dict:
        """Generate a single timetable attempt"""
        # Initialize data structures
        timetable = {day: {slot: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} 
                         for slot in range(19)} for day in range(len(days))}
        professor_schedule = defaultdict(lambda: {day: set() for day in range(len(days))})
        unscheduled_components = set()
        
        # Priority-based scheduling
        courses_by_priority = self._categorize_courses_by_priority(df)
        
        for priority, courses in courses_by_priority.items():
            if enable_parallel and len(courses) > 10:
                self._schedule_courses_parallel(courses, timetable, professor_schedule, 
                                              unscheduled_components, rooms)
            else:
                self._schedule_courses_sequential(courses, timetable, professor_schedule, 
                                                unscheduled_components, rooms)
        
        return {
            'timetable': timetable,
            'professor_schedule': professor_schedule,
            'unscheduled_components': unscheduled_components
        }
    
    def _categorize_courses_by_priority(self, df: pd.DataFrame) -> dict:
        """Categorize courses by scheduling priority"""
        priority_order = self.config.config['scheduling']['priority_order']
        courses_by_priority = {priority: [] for priority in priority_order}
        
        for _, course in df.iterrows():
            code = str(course['Course Code'])
            
            if code.startswith('B') and '-' in code:
                courses_by_priority['basket_electives'].append(course)
            elif any(dept in code for dept in ['CS', 'EC', 'MA', 'PH']):
                courses_by_priority['core_courses'].append(course)
            elif course.get('L', 0) > 0:
                courses_by_priority['regular_electives'].append(course)
            elif course.get('T', 0) > 0:
                courses_by_priority['tutorials'].append(course)
            elif course.get('P', 0) > 0:
                courses_by_priority['labs'].append(course)
            else:
                courses_by_priority['self_study'].append(course)
        
        return courses_by_priority
    
    def _schedule_courses_parallel(self, courses: List[pd.Series], timetable: dict, 
                                 professor_schedule: dict, unscheduled_components: set, 
                                 rooms: dict):
        """Schedule courses in parallel"""
        max_workers = self.config.config['optimization']['max_workers']
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = []
            for course in courses:
                future = executor.submit(self._schedule_single_course, course, 
                                       timetable, professor_schedule, rooms)
                futures.append(future)
            
            for future in as_completed(futures):
                try:
                    result = future.result()
                    if result:
                        unscheduled_components.add(result)
                except Exception as e:
                    logging.error(f"Error scheduling course in parallel: {e}")
    
    def _schedule_courses_sequential(self, courses: List[pd.Series], timetable: dict,
                                   professor_schedule: dict, unscheduled_components: set,
                                   rooms: dict):
        """Schedule courses sequentially"""
        for course in courses:
            result = self._schedule_single_course(course, timetable, professor_schedule, rooms)
            if result:
                unscheduled_components.add(result)
    
    def _schedule_single_course(self, course: pd.Series, timetable: dict, 
                              professor_schedule: dict, rooms: dict) -> Optional[object]:
        """Schedule a single course (placeholder implementation)"""
        # This would contain the actual scheduling logic
        # For now, return None (scheduled successfully)
        return None
    
    def _calculate_schedule_score(self, result: dict) -> int:
        """Calculate a score for the schedule quality (lower is better)"""
        if not result:
            return float('inf')
        
        unscheduled_count = len(result.get('unscheduled_components', []))
        conflicts = self._count_conflicts(result.get('timetable', {}))
        
        return unscheduled_count + conflicts
    
    def _count_conflicts(self, timetable: dict) -> int:
        """Count conflicts in the timetable"""
        conflicts = 0
        # Implementation would check for overlapping schedules, etc.
        return conflicts

class EnhancedTimetableGenerator:
    """Main timetable generator with enhanced features"""
    
    def __init__(self, config_file: str = 'config.json'):
        self.config = TimetableConfig(config_file)
        self.scheduler = OptimizedScheduler(self.config)
        self.setup_constants()
    
    def setup_constants(self):
        """Setup constants from configuration"""
        settings = self.config.config
        
        # Time settings
        self.DAYS = settings['timetable_settings']['days']
        self.START_TIME = dt_time(*map(int, settings['timetable_settings']['start_time'].split(':')))
        self.END_TIME = dt_time(*map(int, settings['timetable_settings']['end_time'].split(':')))
        
        # Duration settings
        durations = settings['course_durations']
        self.LECTURE_DURATION = durations['lecture_duration_slots']
        self.LAB_DURATION = durations['lab_duration_slots']
        self.TUTORIAL_DURATION = durations['tutorial_duration_slots']
        self.SELF_STUDY_DURATION = durations['self_study_duration_slots']
        
        # Basket settings
        if settings.get('basket_electives', {}).get('enabled', True):
            self.BASKET_SLOTS = settings['basket_electives']['fixed_slots']
            self.basket_group_colors = settings['basket_electives']['colors']
        
        # Colors
        self.COLOR_PALETTE = settings['colors']['color_palette']
        self.activity_colors = settings['colors']['activity_colors']
    
    def load_data(self) -> Tuple[pd.DataFrame, dict]:
        """Load course and room data with fallback options"""
        files = self.config.config['files']
        
        # Load courses
        try:
            df = pd.read_csv(files['courses_file'])
        except FileNotFoundError:
            try:
                df = pd.read_csv(files['fallback_courses'])
            except FileNotFoundError:
                logging.warning("No course files found, using default data")
                df = self._get_default_courses()
        
        # Load rooms
        try:
            rooms = self._load_rooms(files['rooms_file'])
        except FileNotFoundError:
            try:
                rooms = self._load_rooms(files['fallback_rooms'])
            except FileNotFoundError:
                logging.warning("No room files found, using default rooms")
                rooms = self._get_default_rooms()
        
        return df, rooms
    
    def _load_rooms(self, filename: str) -> dict:
        """Load room data from CSV"""
        rooms = {}
        with open(filename, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rooms[row['id']] = {
                    'capacity': int(row['capacity']),
                    'type': row['type'],
                    'roomNumber': row['roomNumber'],
                    'schedule': {day: set() for day in range(len(self.DAYS))}
                }
        return rooms
    
    def _get_default_courses(self) -> pd.DataFrame:
        """Return default course data"""
        return pd.DataFrame([
            {'Department': 'CSE', 'Semester': 3, 'Course Code': 'CS101', 'Course Name': 'Intro to CS', 
             'Faculty': 'Prof A', 'L': 3, 'T': 1, 'P': 2, 'S': 0, 'total_students': 140},
            {'Department': 'ECE', 'Semester': 3, 'Course Code': 'EC101', 'Course Name': 'Electronics', 
             'Faculty': 'Prof B', 'L': 3, 'T': 0, 'P': 2, 'S': 0, 'total_students': 70}
        ])
    
    def _get_default_rooms(self) -> dict:
        """Return default room data"""
        return {
            'R1': {'capacity': 70, 'type': 'LECTURE_ROOM', 'roomNumber': 'R101', 
                   'schedule': {day: set() for day in range(len(self.DAYS))}},
            'L1': {'capacity': 35, 'type': 'COMPUTER_LAB', 'roomNumber': 'L101', 
                   'schedule': {day: set() for day in range(len(self.DAYS))}}
        }
    
    def generate_timetable(self):
        """Main timetable generation function"""
        logging.info("Starting enhanced timetable generation")
        
        # Load data
        df, rooms = self.load_data()
        
        # Generate timetable with optimization
        result = self.scheduler.generate_timetable_optimized(df, rooms)
        
        if result:
            # Generate Excel output
            self._create_excel_output(result, df)
            
            # Log performance statistics
            self._log_performance_stats()
            
            # Generate conflict resolution suggestions
            if result.get('unscheduled_components'):
                suggestions = self.scheduler.conflict_resolver.suggest_conflict_resolution(
                    list(result['unscheduled_components'])
                )
                self._save_conflict_suggestions(suggestions)
        
        logging.info("Timetable generation completed")
    
    def _create_excel_output(self, result: dict, df: pd.DataFrame):
        """Create Excel output file from the generated timetable result"""
        settings = self.config.config
        output_file = settings['files']['output_file']

        # Ensure output directory exists
        out_dir = os.path.dirname(output_file)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        # Generate time slots based on config
        start_h, start_m = map(int, settings['timetable_settings']['start_time'].split(':'))
        end_h, end_m = map(int, settings['timetable_settings']['end_time'].split(':'))
        slot_minutes = settings['timetable_settings']['slot_duration_minutes']

        def generate_time_slots():
            slots = []
            cur_h, cur_m = start_h, start_m
            while (cur_h, cur_m) < (end_h, end_m):
                next_m = cur_m + slot_minutes
                next_h = cur_h + next_m // 60
                next_m = next_m % 60
                slots.append((f"{cur_h:02d}:{cur_m:02d}", f"{next_h:02d}:{next_m:02d}"))
                cur_h, cur_m = next_h, next_m
            return slots

        time_slots = generate_time_slots()
        days = settings['timetable_settings']['days']

        # Build workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Timetable"

        # Header row
        header = ['Day'] + [f"{s}-{e}" for s, e in time_slots]
        ws.append(header)

        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Colors
        activity_colors = self.config.config['colors']['activity_colors']

        timetable = result.get('timetable', {})

        # Write each day row
        merge_ranges = []
        for day_idx, day_name in enumerate(days):
            row_num = day_idx + 2
            ws.append([day_name])

            for slot_idx in range(len(time_slots)):
                cell_value = ''
                cell_fill = None

                slot_data = timetable.get(day_idx, {}).get(slot_idx, {})
                if slot_data and slot_data.get('type'):
                    activity_type = slot_data['type']
                    code = slot_data.get('code', '')
                    classroom = slot_data.get('classroom', '')
                    faculty = slot_data.get('faculty', '')

                    # Determine duration by extending while same activity appears
                    duration = 1
                    for i in range(slot_idx + 1, min(len(time_slots), slot_idx + 10)):
                        nxt = timetable.get(day_idx, {}).get(i, {})
                        if nxt and nxt.get('type') == activity_type and not nxt.get('code'):
                            duration += 1
                        else:
                            break

                    # Fill and value
                    color_key = 'lecture'
                    if activity_type == 'LAB':
                        color_key = 'lab'
                    elif activity_type == 'TUT':
                        color_key = 'tutorial'
                    elif activity_type == 'SS':
                        color_key = 'self_study'

                    fill_color = activity_colors.get({
                        'LEC': 'lecture', 'LAB': 'lab', 'TUT': 'tutorial', 'SS': 'self_study'
                    }.get(activity_type, 'lecture'), 'E6E6FA')

                    cell_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell_value = f"{code} {activity_type}\n{classroom}\n{faculty}" if code else ''

                    if duration > 1:
                        start_col = get_column_letter(slot_idx + 2)
                        end_col = get_column_letter(slot_idx + duration + 1)
                        merge_ranges.append((f"{start_col}{row_num}:{end_col}{row_num}", cell_fill, row_num, slot_idx + 2))

                cell = ws.cell(row=row_num, column=slot_idx + 2, value=cell_value)
                if cell_fill:
                    cell.fill = cell_fill
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

            # Set border and alignment for day label
            ws.cell(row=row_num, column=1).border = border
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')

        # Perform merges after filling
        for merge_range, fill, row_num, col_start in merge_ranges:
            ws.merge_cells(merge_range)
            merged_cell = ws.cell(row=row_num, column=col_start)
            merged_cell.fill = fill
            merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # Column widths and row heights
        for col_idx in range(1, len(time_slots) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        for row in ws.iter_rows(min_row=2, max_row=len(days) + 1):
            ws.row_dimensions[row[0].row].height = 40

        wb.save(output_file)
        logging.info(f"Excel output created: {output_file}")
    
    def _log_performance_stats(self):
        """Log performance statistics"""
        stats = self.scheduler.performance_stats
        logging.info(f"Performance Statistics:")
        logging.info(f"  Generation time: {stats['generation_time']:.2f} seconds")
        logging.info(f"  Retry attempts: {stats['retry_attempts']}")
        logging.info(f"  Conflicts resolved: {stats['conflicts_resolved']}")
    
    def _save_conflict_suggestions(self, suggestions: dict):
        """Save conflict resolution suggestions to file"""
        with open('conflict_suggestions.json', 'w') as f:
            json.dump(suggestions, f, indent=2)
        logging.info("Conflict suggestions saved to conflict_suggestions.json")

if __name__ == "__main__":
    generator = EnhancedTimetableGenerator()
    generator.generate_timetable()
